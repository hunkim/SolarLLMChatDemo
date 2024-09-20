# from https://docs.streamlit.io/develop/tutorials/llms/build-conversational-apps

import streamlit as st
from langchain_upstage import (
    UpstageLayoutAnalysisLoader,
    UpstageGroundednessCheck,
    ChatUpstage,
)
from langchain_core.output_parsers import StrOutputParser
from langchain.prompts import ChatPromptTemplate

from openpyxl import Workbook

import io
import os
import re
import tempfile
import unicodedata
import zipfile
import hashlib


if "processed_files" not in st.session_state:
    st.session_state.processed_files = set()
if "students_data" not in st.session_state:
    st.session_state.students_data = []


st.title("Solar HW Grader")
st.write(
    "This is Solar SNU HW grader demo. Get your KEY at https://console.upstage.ai/"
)

llm = ChatUpstage(model="solar-pro")

hw_prompt = ChatPromptTemplate.from_messages(
    [
        (
            "system",
            """You are Prof. Solar, very nice and smart, loved by many people. 
            """,
        ),
        (
            "human",
            """For given report, please provide score 1-5 and quick summary of the report and explain your score and provide advice. Format your response as follows:
         Score: [score]
         Summary: [summary]
         Explanation: [explanation]
         Advice: [advice]
         ---
         Student report: {student_report},
         """,
        ),
    ]
)

groundedness_check = UpstageGroundednessCheck()


def get_response(retrieved_docs):
    chain = hw_prompt | llm | StrOutputParser()

    return chain.stream(
        {
            "student_report": retrieved_docs,
        }
    )


def hash_filename(filename):
    # Generate a hash of the filename
    return hashlib.md5(filename.encode("utf-8")).hexdigest() + ".pdf"


def create_excel_grade(students_data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Grades"

    ws["A1"] = "File Name"
    ws["B1"] = "Score"
    ws["C1"] = "Feedback"

    for row, (name, score, feedback) in enumerate(students_data, start=2):
        # Normalize the Korean name to composed form
        normalized_name = unicodedata.normalize("NFC", name)
        ws[f"A{row}"] = normalized_name
        ws[f"B{row}"] = score
        ws[f"C{row}"] = feedback

    return wb


def process_pdf_file(file_path):
    with st.status(f"Document Parsing {file_path}..."):
        layzer = UpstageLayoutAnalysisLoader(file_path, split="page")
        # For improved memory efficiency, consider using the lazy_load method to load documents page by page.
        docs = layzer.load()  # or layzer.lazy_load()

    with st.chat_message("user"):
        st.markdown(f"Grading {file_path}")

    file_name = os.path.basename(file_path)
    student_name_match = re.search(r"^(.*?)(?=\d)", file_name, re.UNICODE)
    student_name = (
        student_name_match.group(1).strip() if student_name_match else "Unknown"
    )

    with st.chat_message("assistant"):
        full_response = ""
        response_placeholder = st.empty()
        for chunk in get_response(docs):
            full_response += chunk
            response_placeholder.markdown(full_response)

        score_match = re.search(r"Score: (\d+)", full_response)
        score = score_match.group(1) if score_match else "N/A"

        return student_name, score, full_response


uploaded_file = st.file_uploader(
    "Choose your `.pdf` or `.zip` file", type=["pdf", "zip"]
)

if (
    uploaded_file is not None
    and uploaded_file.name not in st.session_state.processed_files
):
    with tempfile.TemporaryDirectory() as temp_dir:
        file_path = os.path.join(temp_dir, uploaded_file.name)

        with open(file_path, "wb") as f:
            f.write(uploaded_file.getvalue())

        if uploaded_file.name.endswith(".pdf"):
            student_name, score, feedback = process_pdf_file(file_path)
            st.session_state.students_data.append((student_name, score, feedback))

        elif uploaded_file.name.endswith(".zip"):
            with zipfile.ZipFile(file_path, "r") as z:
                for file_info in z.infolist():
                    if file_info.filename.endswith(
                        ".pdf"
                    ) and not file_info.filename.startswith("__MACOSX/"):
                        try:
                            original_filename = file_info.filename.encode(
                                "cp437"
                            ).decode("utf-8")
                            hashed_name = hash_filename(original_filename)
                            extracted_path = os.path.join(temp_dir, hashed_name)
                            with z.open(file_info) as source, open(
                                extracted_path, "wb"
                            ) as target:
                                target.write(source.read())
                            student_name, score, feedback = process_pdf_file(
                                extracted_path
                            )
                            # Use the original filename for display and grading
                            student_name = os.path.splitext(original_filename)[0]
                            st.session_state.students_data.append(
                                (student_name, score, feedback)
                            )
                        except Exception as e:
                            st.error(f"Error processing file {file_info.filename}: {e}")

        st.session_state.processed_files.add(uploaded_file.name)

if st.session_state.students_data:
    wb = create_excel_grade(st.session_state.students_data)

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    st.download_button(
        label="Download Excel Grades",
        data=excel_buffer,
        file_name="grades.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

if st.session_state.processed_files:
    st.write("Processed files:")
    for file in st.session_state.processed_files:
        st.text(file)

if st.session_state.students_data:
    st.write("Current Grades:")
    for name, score, _ in st.session_state.students_data:
        st.text(f"{name}: {score}")

# Add a button to clear the session state
if st.button("Clear All Data"):
    st.session_state.processed_files.clear()
    st.session_state.students_data.clear()
